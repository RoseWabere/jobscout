"""
modules/excel_export.py — JobScout KE
Generates and updates an Excel workbook tracking all applications.
Auto-formats with colour-coded status, hyperlinks, and summary dashboard.
"""
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config import EXPORTS_DIR

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False
    print("[excel] openpyxl not installed — run: pip install openpyxl")

EXPORT_PATH = EXPORTS_DIR / "JobScout_Applications.xlsx"

# Colour palette
_C = {
    "header_bg": "1C2230",
    "header_fg": "E4DDD0",
    "new":       "1A3A3A",   # dark teal
    "pending":   "3A2E0A",   # dark amber
    "applied":   "0A2A1A",   # dark green
    "skipped":   "1C1C1C",   # near black
    "row_alt":   "1A1D26",
    "row_base":  "161B22",
    "link":      "4EA8DE",
}

_STATUS_FILL = {
    "new":              PatternFill("solid", fgColor="0D3B38") if _OPENPYXL else None,
    "pending_approval": PatternFill("solid", fgColor="3D2E00") if _OPENPYXL else None,
    "applied":          PatternFill("solid", fgColor="0B2D1A") if _OPENPYXL else None,
    "skipped":          PatternFill("solid", fgColor="1A1A1A") if _OPENPYXL else None,
}

_COLS = [
    ("ID",            8),
    ("Job Title",     30),
    ("Company",       22),
    ("Source",        14),
    ("Location",      16),
    ("Match Score",   13),
    ("Status",        16),
    ("Date Found",    14),
    ("Date Applied",  14),
    ("Job URL",       40),
    ("Notes",         30),
]


def export_to_excel(jobs: list[dict]) -> Path:
    """
    Generate / overwrite the Excel workbook with all jobs.
    Returns path to the file.
    """
    if not _OPENPYXL:
        raise RuntimeError("openpyxl not installed")

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Applications ────────────────────────────────────
    ws = wb.active
    ws.title = "Applications"
    ws.sheet_view.showGridLines = False

    # Tab colour
    ws.sheet_properties.tabColor = "2A9D8F"

    # Freeze panes below header
    ws.freeze_panes = "A2"

    # Header row
    header_font = Font(name="Calibri", bold=True, color=_C["header_fg"], size=10)
    header_fill = PatternFill("solid", fgColor=_C["header_bg"])
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin = Side(style="thin", color="2A2A2A")
    border = Border(bottom=Side(style="medium", color="2A9D8F"))

    for col_idx, (col_name, col_w) in enumerate(_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, job in enumerate(jobs, 2):
        score  = job.get("match_score", 0)
        status = job.get("status", "new")
        url    = job.get("url","")

        row_data = [
            job.get("id",""),
            job.get("title",""),
            job.get("company",""),
            job.get("source",""),
            job.get("location",""),
            f"{score}%",
            status.replace("_"," ").title(),
            job.get("scraped_at","")[:10],
            job.get("applied_at","")[:10] if job.get("applied_at") else "",
            url,
            "",
        ]

        fill = _STATUS_FILL.get(status, PatternFill("solid", fgColor=_C["row_base"]))

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.font = Font(name="Calibri", color="D0D8E0", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Colour code score
            if col_idx == 6:
                if score >= 70:
                    cell.font = Font(name="Calibri", bold=True, color="5A9E6F", size=9)
                elif score >= 50:
                    cell.font = Font(name="Calibri", bold=True, color="C9A845", size=9)
                else:
                    cell.font = Font(name="Calibri", color="707878", size=9)

            # Hyperlink for URL column
            if col_idx == 10 and url:
                cell.value = "Open"
                cell.hyperlink = url
                cell.font = Font(name="Calibri", color=_C["link"], underline="single", size=9)

        ws.row_dimensions[row_idx].height = 16

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS))}{len(jobs)+1}"

    # ── Sheet 2: Summary dashboard ───────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_properties.tabColor = "C8872A"
    ws2.sheet_view.showGridLines = False

    statuses = ["new", "pending_approval", "applied", "skipped"]
    counts   = {s: sum(1 for j in jobs if j.get("status")==s) for s in statuses}
    avg_score = sum(j.get("match_score",0) for j in jobs) / max(len(jobs),1)
    applied_list = [j for j in jobs if j.get("status")=="applied"]

    summary_rows = [
        ["JobScout KE — Application Summary", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["", ""],
        ["Total tracked", len(jobs)],
        ["New / unreviewed", counts["new"]],
        ["Pending approval", counts["pending_approval"]],
        ["Applied", counts["applied"]],
        ["Skipped", counts["skipped"]],
        ["Average match score", f"{avg_score:.1f}%"],
        ["", ""],
        ["Applications this week", sum(
            1 for j in applied_list
            if j.get("applied_at","") >= (datetime.now().strftime("%Y-%m-%d")[:8] + "01")
        )],
    ]

    hdr_font2 = Font(name="Calibri", bold=True, color="E4DDD0", size=11)
    body_font2 = Font(name="Calibri", color="B0A898", size=10)

    for r, (label, value) in enumerate(summary_rows, 1):
        c1 = ws2.cell(row=r, column=1, value=label)
        c2 = ws2.cell(row=r, column=2, value=value)
        if r == 1:
            c1.font = Font(name="Calibri", bold=True, color="2A9D8F", size=13)
        else:
            c1.font = hdr_font2
            c2.font = body_font2

    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 20
    ws2.sheet_view.showGridLines = False

    # ── Sheet 3: High Matches ────────────────────────────────────────
    top = sorted([j for j in jobs if j.get("match_score",0) >= 60],
                 key=lambda x: x.get("match_score",0), reverse=True)
    if top:
        ws3 = wb.create_sheet("Top Matches (60%+)")
        ws3.sheet_properties.tabColor = "5A9E6F"
        ws3.sheet_view.showGridLines = False

        hdr3 = Font(name="Calibri", bold=True, color=_C["header_fg"], size=10)
        hdr3_fill = PatternFill("solid", fgColor="0B2D1A")
        cols3 = [("Score","8"),("Title","30"),("Company","22"),("Source","14"),("URL","40")]
        for ci, (cn, cw) in enumerate(cols3, 1):
            cell = ws3.cell(row=1, column=ci, value=cn)
            cell.font = hdr3
            cell.fill = hdr3_fill
            ws3.column_dimensions[get_column_letter(ci)].width = int(cw)

        for ri, j in enumerate(top, 2):
            url = j.get("url","")
            vals = [
                f"{j.get('match_score',0)}%",
                j.get("title",""),
                j.get("company",""),
                j.get("source",""),
                url,
            ]
            fill3 = PatternFill("solid", fgColor="0B2D1A")
            for ci, v in enumerate(vals, 1):
                cell = ws3.cell(row=ri, column=ci, value=v if ci < 5 else "Open")
                cell.fill = fill3
                cell.font = Font(name="Calibri", color="D0D8E0", size=9)
                if ci == 5 and url:
                    cell.hyperlink = url
                    cell.font = Font(name="Calibri", color=_C["link"], underline="single", size=9)

    wb.save(str(EXPORT_PATH))
    return EXPORT_PATH
